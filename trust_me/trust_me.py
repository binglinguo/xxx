import subprocess
import sys
import os
import openpyxl
import sqlite3
import json
import platform
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
import asyncio

from IPython.lib import clipboard
from PyQt5.QtWidgets import (QTextEdit, QVBoxLayout, QWidget, QMessageBox,
                             QApplication, QMainWindow, QLabel, QLineEdit,
                             QPushButton, QProgressBar, QFileDialog, QMenu,
                             QDialog)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QTextCursor, QTextCharFormat, QColor


class IndexingDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("建立索引")
        self.setFixedSize(300, 100)

        layout = QVBoxLayout(self)

        self.status_label = QLabel("正在扫描文件...")
        layout.addWidget(self.status_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        layout.addWidget(self.progress_bar)

    def update_progress(self, value, text=None):
        self.progress_bar.setValue(value)
        if text:
            self.status_label.setText(text)


class Config:
    def __init__(self):
        self.config_file = "config.json"
        self.load_config()

    def load_config(self):
        try:
            with open(self.config_file, 'r') as f:
                self.config = json.load(f)
        except FileNotFoundError:
            self.config = {'last_folder': ''}

    def save_config(self):
        with open(self.config_file, 'w') as f:
            json.dump(self.config, f)

    def get_last_folder(self):
        return self.config.get('last_folder', '')

    def set_last_folder(self, folder):
        self.config['last_folder'] = folder
        self.save_config()


class DatabaseManager:
    def __init__(self):
        self.db_path = 'file_index.db'
        self.create_tables()

    def get_connection(self):
        # 为每个线程创建新的数据库连接
        return sqlite3.connect(self.db_path)

    def create_tables(self):
        with self.get_connection() as conn:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS files (
                    id INTEGER PRIMARY KEY,
                    file_path TEXT UNIQUE,
                    last_modified TIMESTAMP,
                    last_indexed TIMESTAMP
                )
            ''')
            conn.execute('''
                CREATE TABLE IF NOT EXISTS content_index (
                    id INTEGER PRIMARY KEY,
                    file_id INTEGER,
                    sheet_name TEXT,
                    cell_coordinate TEXT,
                    content TEXT,
                    FOREIGN KEY (file_id) REFERENCES files(id)
                )
            ''')

    def update_index(self, folder_path, progress_callback=None):
        current_time = datetime.now()

        # 首先扫描所有文件
        excel_files = []
        for dirpath, _, filenames in os.walk(folder_path):
            for filename in filenames:
                if filename.endswith(('.xlsx', '.xlsm')) and not filename.startswith('~$'):
                    excel_files.append(os.path.join(dirpath, filename))

        # 在这个线程中创建新的数据库连接
        conn = self.get_connection()
        try:
            total_files = len(excel_files)
            for i, file_path in enumerate(excel_files):
                try:
                    file_modified_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                    if self.needs_update(conn, file_path, file_modified_time):
                        if progress_callback:
                            progress_callback(
                                int((i + 1) * 100 / total_files),
                                f"正在索引: {os.path.basename(file_path)}"
                            )
                        self.index_file(conn, file_path, file_modified_time, current_time)
                except Exception as e:
                    print(f"处理文件 {file_path} 时出错: {e}")

            if progress_callback:
                progress_callback(100, "索引完成")
        finally:
            conn.close()

    def needs_update(self, conn, file_path, modified_time):
        cursor = conn.cursor()
        cursor.execute('''
            SELECT last_modified FROM files WHERE file_path = ?
        ''', (file_path,))
        result = cursor.fetchone()
        return result is None or result[0] < modified_time

    def index_file(self, conn, file_path, modified_time, current_time):
        try:
            wb = openpyxl.load_workbook(filename=file_path,
                                        data_only=True,
                                        keep_vba=False,
                                        keep_links=False,
                                        read_only=True)

            cursor = conn.cursor()
            cursor.execute('BEGIN TRANSACTION')
            try:
                cursor.execute('DELETE FROM files WHERE file_path = ?', (file_path,))
                cursor.execute('DELETE FROM content_index WHERE file_id IN (SELECT id FROM files WHERE file_path = ?)', (file_path,))

                cursor.execute('''
                    INSERT INTO files (file_path, last_modified, last_indexed)
                    VALUES (?, ?, ?)
                ''', (file_path, modified_time, current_time))
                file_id = cursor.lastrowid

                content_values = []

                for sheet_name in wb.sheetnames:
                    try:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value:
                                    content_values.append((
                                        file_id,
                                        sheet_name,
                                        cell.coordinate,
                                        str(cell.value)
                                    ))
                    except Exception as sheet_error:
                        print(f"处理工作表 {sheet_name} 时出错: {sheet_error}")
                        continue

                cursor.executemany('''
                    INSERT INTO content_index 
                    (file_id, sheet_name, cell_coordinate, content)
                    VALUES (?, ?, ?, ?)
                ''', content_values)

                conn.commit()

            except Exception as e:
                conn.rollback()
                raise e
            finally:
                wb.close()

        except Exception as e:
            print(f"索引文件 {file_path} 时出错: {e}")
            with open('index_errors.log', 'a', encoding='utf-8') as f:
                f.write(f"{datetime.now()}: 文件 {file_path} 索引失败: {str(e)}\n")


    def search(self, query):
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT f.file_path, ci.sheet_name, ci.cell_coordinate, ci.content
                    FROM content_index ci
                    JOIN files f ON ci.file_id = f.id
                    WHERE ci.content LIKE ?
                ''', (f"%{query}%",))
                return cursor.fetchall()
        except sqlite3.Error as e:
            print(f"数据库查询错误: {e}")
            return []


class ClickableTextEdit(QTextEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
        self.setMouseTracking(True)
        self.file_positions = []

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            cursor = self.cursorForPosition(event.pos())
            position = cursor.position()
            for start, end, file_path in self.file_positions:
                if start <= position <= end:
                    open_file(file_path)
                    break
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        cursor = self.cursorForPosition(event.pos())
        position = cursor.position()

        is_over_link = False
        for start, end, _ in self.file_positions:
            if start <= position <= end:
                is_over_link = True
                break

        if is_over_link:
            self.viewport().setCursor(Qt.PointingHandCursor)
        else:
            self.viewport().setCursor(Qt.IBeamCursor)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = Config()
        self.db_manager = DatabaseManager()
        self.executor = ThreadPoolExecutor(max_workers=1)
        self.indexing_thread = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Excel 文件查询工具")
        self.setGeometry(100, 100, 800, 600)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        self.query_id_label = QLabel("查找内容:")
        layout.addWidget(self.query_id_label)

        self.query_id_input = QLineEdit()
        layout.addWidget(self.query_id_input)

        self.folder_button = QPushButton('选择文件夹')
        self.folder_button.clicked.connect(self.open_folder_dialog)
        layout.addWidget(self.folder_button)

        self.folder_label = QLabel()
        layout.addWidget(self.folder_label)

        self.query_button = QPushButton("查询")
        self.query_button.clicked.connect(self.start_query)
        layout.addWidget(self.query_button)

        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        self.result_text = ClickableTextEdit()
        self.result_text.setContextMenuPolicy(Qt.CustomContextMenu)
        self.result_text.customContextMenuRequested.connect(self.show_context_menu)
        layout.addWidget(self.result_text)

        last_folder = self.config.get_last_folder()
        if last_folder:
            self.folder_label.setText(last_folder)

    def open_folder_dialog(self):
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "选择文件夹",
            self.config.get_last_folder()
        )
        if folder_path:
            self.folder_label.setText(folder_path)
            self.config.set_last_folder(folder_path)

            # 显示索引进度对话框
            indexing_dialog = IndexingDialog(self)
            indexing_dialog.show()

            # 创建一个新线程来执行索引操作
            class IndexingThread(QThread):
                finished = pyqtSignal()

                def __init__(self, db_manager, folder_path, progress_callback):
                    super().__init__()
                    self.db_manager = db_manager
                    self.folder_path = folder_path
                    self.progress_callback = progress_callback

                def run(self):
                    self.db_manager.update_index(self.folder_path, self.progress_callback)
                    self.finished.emit()

            # 创建并启动索引线程
            self.indexing_thread = IndexingThread(
                self.db_manager,
                folder_path,
                indexing_dialog.update_progress
            )

            # 当索引完成时关闭对话框
            self.indexing_thread.finished.connect(indexing_dialog.accept)
            self.indexing_thread.finished.connect(
                lambda: QMessageBox.information(self, "完成", "索引建立完成！")
            )

            # 启动线程
            self.indexing_thread.start()

    def start_query(self):
        query = self.query_id_input.text()
        folder_path = self.folder_label.text()

        if not query or not folder_path:
            QMessageBox.warning(self, "警告", "请输入查询内容并选择文件夹")
            return

        self.result_text.clear()
        self.result_text.file_positions = []
        self.progress_bar.setValue(0)

        results = self.db_manager.search(query)

        link_format = QTextCharFormat()
        link_format.setForeground(QColor('blue'))
        link_format.setUnderlineStyle(QTextCharFormat.SingleUnderline)

        cursor = self.result_text.textCursor()

        for result in results:
            file_path, sheet_name, cell_coordinate, content = result

            start_pos = cursor.position()
            cursor.insertText(file_path, link_format)
            end_pos = cursor.position()

            self.result_text.file_positions.append((start_pos, end_pos, file_path))
            cursor.insertText(f" (工作表: {sheet_name}, 位置: {cell_coordinate}, 内容: {content})\n")

        self.progress_bar.setValue(100)

    def show_context_menu(self, position):
        menu = QMenu()
        menu.exec_(self.result_text.mapToGlobal(position))

    def extract_file_path(self, text):
        start = text.find("'") + 1
        end = text.find("'", start)
        if start > 0 and end > start:
            return text[start:end]
        return None


def open_file(file_path):
    try:
        if platform.system() == 'Windows':
            os.startfile(file_path)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', file_path])
        elif platform.system() == 'Linux':
            subprocess.run(['xdg-open', file_path])
    except Exception as e:
        QMessageBox.warning(None, "错误", f"无法打开文件: {str(e)}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
